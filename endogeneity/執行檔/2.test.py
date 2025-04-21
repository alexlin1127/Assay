import pandas as pd
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor
from openpyxl import Workbook

path1 = '/Users/alex/Desktop/Assay data/analysis data/part 3情緒分析/5. variable data for all.xlsx'  
path2 = '/Users/alex/Desktop/Assay data/analysis data/part 1 事件研究法/CAR windows data.xlsx' 
output_path = '/Users/alex/Desktop/Assay data/analysis data/analysis_results with IMR.xlsx' 

sheets = ['BiLSTM Label', 'BERT Label', 'FinBERT Label', 'FinBERT_ESGCalls Label', 'LM label']

df2 = pd.read_excel(path2)
# y = df2['car[-4 +4]']  
y = df2['car[+0 +2]']
# y = df2['car[+0 +4]']  
# y = df2['car[+0 +6]']  



wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"

for sheet in sheets:
    df1 = pd.read_excel(path1, sheet_name=sheet)

    X = df1[['Net_tone', 'Sentiment_power', 'Vagueness', 'FLS', 'ROA_t-1', 'market cap', 'IMR']]
    # X = df1[['Net_tone', 'Sentiment_power', 'Vagueness', 'Temporality', 'log_Firm_size', 'ROA_t-1', 'IMR']]

    # X = sm.add_constant(X)

    model = sm.OLS(y, X)
    results = model.fit()

    ws = wb.create_sheet(title=f"{sheet}")
    summary_str = results.summary().as_text()
    print('\n\n', f'---- {sheet} result ----','\n')
    print(summary_str, '\n')
    # for row in summary_str.split('\n'):
    #     ws.append([row])

    # correlation_matrix = X.corr()
    # print(correlation_matrix, '\n')

    # vif_data = pd.DataFrame()
    # vif_data["variable"] = X.columns
    # vif_data["VIF"] = [variance_inflation_factor(X.values, i) for i in range(X.shape[1])]

    # print('\n\n', f'---- {sheet} VIF ----','\n')
    # print(vif_data, '\n\n\n')

    # ws_corr = wb.create_sheet(title=f"{sheet}_correlation")
    # for r_idx, row in enumerate(correlation_matrix.itertuples(), 1):
    #     for c_idx, value in enumerate(row, 1):
    #         ws_corr.cell(row=r_idx, column=c_idx, value=value)

# wb.save(output_path)

# print(f"Analysis results have been saved to {output_path}")
